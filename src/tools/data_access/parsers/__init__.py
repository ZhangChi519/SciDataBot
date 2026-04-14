"""Data format parsers for various file formats."""
import io
import json
import csv
import zipfile
import tarfile
import gzip
from pathlib import Path
from typing import Any, Optional, Union
from abc import ABC, abstractmethod

from src.tools.base import Tool, ToolResult, ToolCategory


class DataParser(ABC):
    """Base class for data parsers."""

    @abstractmethod
    async def parse(self, content: Union[str, bytes], **kwargs) -> dict:
        """Parse data and return structured result."""
        pass

    @abstractmethod
    async def serialize(self, data: dict, **kwargs) -> bytes:
        """Serialize data to format."""
        pass


class CSVParser(DataParser):
    """CSV parser with flexible options."""

    async def parse(self, content: Union[str, bytes], **kwargs) -> dict:
        """Parse CSV content.

        Args:
            content: CSV content as string or bytes
            **kwargs: csv.DictReader options (delimiter, quotechar, etc.)
        """
        delimiter = kwargs.get("delimiter", ",")
        quotechar = kwargs.get("quotechar", '"')
        skip_rows = kwargs.get("skip_rows", 0)
        max_rows = kwargs.get("max_rows", None)

        if isinstance(content, bytes):
            content = content.decode("utf-8")

        lines = content.splitlines()

        # Skip rows
        if skip_rows > 0:
            lines = lines[skip_rows:]

        reader = csv.reader(lines, delimiter=delimiter, quotechar=quotechar)

        rows = []
        headers = None

        for i, row in enumerate(reader):
            if max_rows and i >= max_rows:
                break

            if i == 0:
                headers = row
            else:
                if headers:
                    rows.append(dict(zip(headers, row)))
                else:
                    rows.append(row)

        return {
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
            "column_count": len(headers) if headers else 0,
        }

    async def serialize(self, data: dict, **kwargs) -> bytes:
        """Serialize data to CSV."""
        delimiter = kwargs.get("delimiter", ",")
        headers = data.get("headers", [])
        rows = data.get("rows", [])

        output = io.StringIO()
        writer = csv.writer(output, delimiter=delimiter)

        if headers:
            writer.writerow(headers)

        for row in rows:
            if isinstance(row, dict):
                writer.writerow([row.get(h, "") for h in headers])
            else:
                writer.writerow(row)

        return output.getvalue().encode("utf-8")


class JSONParser(DataParser):
    """JSON parser with support for JSONL and JSON Lines."""

    async def parse(self, content: Union[str, bytes], **kwargs) -> dict:
        """Parse JSON content.

        Args:
            content: JSON content
            **kwargs: Options like json_lines (bool)
        """
        json_lines = kwargs.get("json_lines", False)

        if isinstance(content, bytes):
            content = content.decode("utf-8")

        if json_lines:
            # Parse as JSON Lines (one JSON object per line)
            objects = []
            for line in content.strip().splitlines():
                if line.strip():
                    objects.append(json.loads(line))
            return {
                "type": "jsonl",
                "objects": objects,
                "count": len(objects),
            }
        else:
            # Parse as regular JSON
            data = json.loads(content)

            # Handle JSON with root array
            if isinstance(data, list):
                return {
                    "type": "array",
                    "data": data,
                    "count": len(data),
                }
            else:
                return {
                    "type": "object",
                    "data": data,
                    "keys": list(data.keys()) if isinstance(data, dict) else [],
                }

    async def serialize(self, data: dict, **kwargs) -> bytes:
        """Serialize data to JSON."""
        json_lines = kwargs.get("json_lines", False)
        indent = kwargs.get("indent", 2)

        if json_lines:
            lines = [json.dumps(obj) for obj in data.get("objects", [])]
            return "\n".join(lines).encode("utf-8")
        else:
            return json.dumps(data, indent=indent).encode("utf-8")


class ExcelParser(DataParser):
    """Excel parser for .xlsx and .xls files."""

    async def parse(self, content: Union[str, bytes], **kwargs) -> dict:
        """Parse Excel content.

        Args:
            content: Excel file content as bytes
            **kwargs: Options like sheet_name, max_rows
        """
        try:
            import openpyxl
        except ImportError:
            return {"error": "openpyxl not installed. Install with: pip install openpyxl"}

        sheet_name = kwargs.get("sheet_name", 0)
        max_rows = kwargs.get("max_rows")

        # Load from bytes
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        if isinstance(sheet_name, int):
            if sheet_name < len(wb.sheetnames):
                ws = wb.worksheets[sheet_name]
            else:
                return {"error": f"Sheet index {sheet_name} out of range"}
        else:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        # Get headers
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        # Get rows
        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if max_rows and i > max_rows + 1:
                break
            rows.append(dict(zip(headers, row)) if headers else list(row))

        return {
            "sheet_name": ws.title,
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
            "column_count": len(headers),
        }

    async def serialize(self, data: dict, **kwargs) -> bytes:
        """Serialize data to Excel."""
        try:
            import openpyxl
        except ImportError:
            return b""

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = kwargs.get("sheet_name", "Sheet1")

        headers = data.get("headers", [])
        rows = data.get("rows", [])

        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(1, col, header)

        # Write rows
        for row_idx, row in enumerate(rows, 2):
            if isinstance(row, dict):
                for col, header in enumerate(headers, 1):
                    ws.cell(row_idx, col, row.get(header))
            else:
                for col, value in enumerate(row, 1):
                    ws.cell(row_idx, col, value)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


class ParquetParser(DataParser):
    """Apache Parquet parser."""

    async def parse(self, content: Union[str, bytes], **kwargs) -> dict:
        """Parse Parquet content."""
        try:
            import pyarrow.parquet as pq
        except ImportError:
            return {"error": "pyarrow not installed. Install with: pip install pyarrow"}

        table = pq.read_table(io.BytesIO(content))

        df = table.to_pandas()

        return {
            "columns": list(df.columns),
            "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
            "rows": df.to_dict("records"),
            "row_count": len(df),
            "column_count": len(df.columns),
        }

    async def serialize(self, data: dict, **kwargs) -> bytes:
        """Serialize data to Parquet."""
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except ImportError:
            return b""

        rows = data.get("rows", [])
        columns = data.get("columns", [])

        if rows and isinstance(rows[0], dict):
            import pandas as pd
            df = pd.DataFrame(rows)
            if columns:
                df = df[columns]
        else:
            import pandas as pd
            df = pd.DataFrame(rows, columns=columns)

        table = pa.Table.from_pandas(df)
        output = io.BytesIO()
        pq.write_table(table, output)
        return output.getvalue()


class NetCDFParser(DataParser):
    """NetCDF parser for scientific data."""

    async def parse(self, content: Union[str, bytes], **kwargs) -> dict:
        """Parse NetCDF content."""
        try:
            import netCDF4
        except ImportError:
            return {"error": "netCDF4 not installed. Install with: pip install netCDF4"}

        dataset = netCDF4.Dataset(
            "memory",
            memory=content if isinstance(content, bytes) else content.encode()
        )

        result = {
            "dimensions": {},
            "variables": {},
            "attributes": {},
        }

        # Dimensions
        for dim_name in dataset.dimensions:
            dim = dataset.dimensions[dim_name]
            result["dimensions"][dim_name] = {
                "size": len(dim),
                "unlimited": dim.isunlimited(),
            }

        # Variables
        for var_name in dataset.variables:
            var = dataset.variables[var_name]
            result["variables"][var_name] = {
                "dimensions": var.dimensions,
                "dtype": str(var.dtype),
                "shape": var.shape,
                "attributes": dict(var.__dict__),
            }

        # Global attributes
        result["attributes"] = dict(dataset.__dict__)

        # Optionally load data
        if kwargs.get("load_data", False):
            for var_name in dataset.variables:
                var = dataset.variables[var_name]
                result["variables"][var_name]["data"] = var[:].tolist()

        dataset.close()
        return result

    async def serialize(self, data: dict, **kwargs) -> bytes:
        """Serialize data to NetCDF."""
        try:
            import netCDF4
        except ImportError:
            return b""

        output = io.BytesIO()

        with netCDF4.Dataset("memory", mode="w", memory=output) as ds:
            # Set dimensions
            for dim_name, dim_info in data.get("dimensions", {}).items():
                size = dim_info.get("size", 1)
                unlimited = dim_info.get("unlimited", False)
                ds.createDimension(dim_name, size if not unlimited else None)

            # Set global attributes
            for attr_name, attr_value in data.get("attributes", {}).items():
                setattr(ds, attr_name, attr_value)

            # Create variables
            for var_name, var_info in data.get("variables", {}).items():
                dims = var_info.get("dimensions", ())
                dtype = var_info.get("dtype", "f8")
                var = ds.createVariable(var_name, dtype, dims)

                # Set attributes
                for attr_name, attr_value in var_info.get("attributes", {}).items():
                    setattr(var, attr_name, attr_value)

                # Set data
                if "data" in var_info:
                    var[:] = var_info["data"]

        return output.getvalue()


class HDF5Parser(DataParser):
    """HDF5 parser for hierarchical data."""

    async def parse(self, content: Union[str, bytes], **kwargs) -> dict:
        """Parse HDF5 content."""
        try:
            import h5py
        except ImportError:
            return {"error": "h5py not installed. Install with: pip install h5py"}

        result = {}

        def parse_group(group, prefix=""):
            output = {}
            for key in group.keys():
                item = group[key]
                item_path = f"{prefix}/{key}" if prefix else key

                if isinstance(item, h5py.Dataset):
                    output[key] = {
                        "type": "dataset",
                        "shape": item.shape,
                        "dtype": str(item.dtype),
                    }
                    if kwargs.get("load_data", False):
                        output[key]["data"] = item[:].tolist()
                elif isinstance(item, h5py.Group):
                    output[key] = {
                        "type": "group",
                        "children": parse_group(item, item_path),
                    }
            return output

        with h5py.File(io.BytesIO(content), "r") as f:
            result = parse_group(f)

        return result

    async def serialize(self, data: dict, **kwargs) -> bytes:
        """Serialize data to HDF5."""
        try:
            import h5py
        except ImportError:
            return b""

        output = io.BytesIO()

        def create_objects(group, obj_data):
            for key, value in obj_data.items():
                if isinstance(value, dict):
                    if value.get("type") == "group":
                        subgroup = group.create_group(key)
                        if "children" in value:
                            create_objects(subgroup, value["children"])
                    elif value.get("type") == "dataset":
                        if "data" in value:
                            group.create_dataset(
                                key,
                                data=value["data"],
                                dtype=value.get("dtype", "f8")
                            )

        with h5py.File(output, "w") as f:
            create_objects(f, data)

        return output.getvalue()


class DataFormatTool(Tool):
    """Tool for parsing and converting various data formats."""

    def __init__(self):
        super().__init__(
            name="data_format",
            description="Parse and convert data formats (CSV, JSON, Excel, Parquet, NetCDF, HDF5)",
            category=ToolCategory.DATA_ACCESS,
        )
        self.parsers = {
            "csv": CSVParser(),
            "json": JSONParser(),
            "xlsx": ExcelParser(),
            "xls": ExcelParser(),
            "parquet": ParquetParser(),
            "nc": NetCDFParser(),
            "hdf5": HDF5Parser(),
            "h5": HDF5Parser(),
        }

    async def execute(self, operation: str, **kwargs) -> ToolResult:
        """Execute data format operation."""
        try:
            if operation == "parse":
                return await self._parse(
                    kwargs.get("content"),
                    kwargs.get("format"),
                    kwargs.get("options", {}),
                )
            elif operation == "convert":
                return await self._convert(
                    kwargs.get("content"),
                    kwargs.get("source_format"),
                    kwargs.get("target_format"),
                    kwargs.get("options", {}),
                )
            elif operation == "detect":
                return await self._detect_format(kwargs.get("content"))
            else:
                return ToolResult(success=False, error=f"Unknown operation: {operation}")
        except Exception as e:
            return ToolResult(success=False, error=str(e))

    async def _parse(self, content: Any, format: str, options: dict) -> ToolResult:
        """Parse content in specified format."""
        if not content:
            return ToolResult(success=False, error="Content is required")

        format_lower = format.lower()
        if format_lower not in self.parsers:
            return ToolResult(
                success=False,
                error=f"Unsupported format: {format}. Supported: {list(self.parsers.keys())}",
            )

        parser = self.parsers[format_lower]
        result = await parser.parse(content, **options)

        if "error" in result:
            return ToolResult(success=False, error=result["error"])

        return ToolResult(success=True, data=result)

    async def _convert(
        self,
        content: Any,
        source_format: str,
        target_format: str,
        options: dict,
    ) -> ToolResult:
        """Convert from one format to another."""
        # Parse source
        parse_result = await self._parse(content, source_format, options)
        if not parse_result.success:
            return parse_result

        # Serialize to target
        target_lower = target_format.lower()
        if target_lower not in self.parsers:
            return ToolResult(
                success=False,
                error=f"Unsupported target format: {target_format}",
            )

        parser = self.parsers[target_lower]
        serialized = await parser.serialize(parse_result.data, **options)

        return ToolResult(
            success=True,
            data={
                "format": target_format,
                "content": serialized.decode("utf-8", errors="ignore") if len(serialized) < 10000 else "[binary data]",
                "size": len(serialized),
                "raw": serialized,
            },
        )

    async def _detect_format(self, content: Any) -> ToolResult:
        """Detect data format from content."""
        if isinstance(content, bytes):
            # Check for binary formats
            if content[:4] == b"PK\x03\x04":  # ZIP/Excel
                return ToolResult(success=True, data={"format": "xlsx", "confidence": 0.9})
            if content[:4] == b"CDF\x02":  # NetCDF
                return ToolResult(success=True, data={"format": "nc", "confidence": 0.9})
            if content[:4] == b"\x89HDF":  # HDF5
                return ToolResult(success=True, data={"format": "hdf5", "confidence": 0.9})

            # Try to decode as text
            try:
                content = content.decode("utf-8")
            except:
                return ToolResult(success=True, data={"format": "unknown", "confidence": 0.0})

        if isinstance(content, str):
            content = content.strip()

            # JSON
            if (content.startswith("{") and content.endswith("}")) or (
                content.startswith("[") and content.endswith("]")
            ):
                try:
                    json.loads(content)
                    return ToolResult(success=True, data={"format": "json", "confidence": 0.9})
                except:
                    pass

            # JSON Lines
            if "\n" in content and all(
                line.startswith("{") or line == ""
                for line in content.splitlines()
                if line.strip()
            ):
                return ToolResult(success=True, data={"format": "jsonl", "confidence": 0.8})

            # CSV (comma or semicolon separated with consistent columns)
            lines = content.splitlines()
            if len(lines) > 1:
                delimiter = "," if "," in lines[0] else ";" if ";" in lines[0] else "\t"
                first_row = lines[0].split(delimiter)
                if len(first_row) > 1:
                    consistent = all(
                        len(line.split(delimiter)) == len(first_row) for line in lines[:5]
                    )
                    if consistent:
                        return ToolResult(
                            success=True,
                            data={"format": "csv", "confidence": 0.7, "delimiter": delimiter},
                        )

        return ToolResult(success=True, data={"format": "unknown", "confidence": 0.0})
